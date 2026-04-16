"""
FerreRAP — API Flask + Supabase
IS2 · UCP · 2026
"""

import io
from datetime import datetime

from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS

from database import supabase, MARGEN_GANANCIA
from models import (
    Alerta, OrdenReposicion,
    GeneradorReporte, ReporteReposicion, ReporteStockActual,
)

# ── PDF ────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# ── Excel ──────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────

app = Flask(__name__)
CORS(app)

USUARIOS = {
    "admin":    {"password": "admin123",    "rol": "Administrador"},
    "empleado": {"password": "empleado123", "rol": "Empleado"},
}

# ── Patron Observer: instanciar observadores con conexion a DB ──
alerta_obs = Alerta(supabase)
orden_obs  = OrdenReposicion(supabase)
observadores = [alerta_obs, orden_obs]


def notificar_bajo_stock(producto):
    """Dispara el patron Observer: notifica a todos los observadores."""
    notificaciones = []
    for obs in observadores:
        resultado = obs.actualizar(producto)
        if resultado:
            notificaciones.append(resultado)
    return notificaciones


def notificar_reposicion(producto):
    """Dispara resolucion del Observer cuando stock se normaliza."""
    notificaciones = []
    for obs in observadores:
        resultado = obs.resolver(producto)
        if resultado:
            if isinstance(resultado, list):
                notificaciones.extend(resultado)
            else:
                notificaciones.append(resultado)
    return notificaciones


# ── Seed ───────────────────────────────────────────────────────
def seed():
    try:
        result = supabase.table("productos").select("id").limit(1).execute()
        if result.data:
            return
    except Exception as e:
        print(f"  ! No se pudo verificar la tabla productos: {e}")
        return

    datos = [
        {"nombre": "Martillo 500g",     "descripcion": "Martillo de carpintero",    "precio_costo": 1000, "precio_venta": 1500, "stock_actual": 8,  "stock_minimo": 5,  "categoria": "Herramientas"},
        {"nombre": "Destornillador Ph", "descripcion": "Phillips punta fina #2",     "precio_costo": 530,  "precio_venta": 800,  "stock_actual": 3,  "stock_minimo": 5,  "categoria": "Herramientas"},
        {"nombre": "Cable 2.5mm x mt",  "descripcion": "Cable unipolar color rojo",  "precio_costo": 230,  "precio_venta": 350,  "stock_actual": 20, "stock_minimo": 10, "categoria": "Electricidad"},
        {"nombre": "Llave de paso 1/2", "descripcion": "Bronce, media pulgada",      "precio_costo": 1470, "precio_venta": 2200, "stock_actual": 2,  "stock_minimo": 4,  "categoria": "Plomeria"},
        {"nombre": "Latex blanco 4L",   "descripcion": "Interior lavable",           "precio_costo": 2530, "precio_venta": 3800, "stock_actual": 12, "stock_minimo": 6,  "categoria": "Pinturas"},
        {"nombre": "Tornillos 4x40",    "descripcion": "Pack x100 unidades",          "precio_costo": 430,  "precio_venta": 650,  "stock_actual": 50, "stock_minimo": 20, "categoria": "Fijaciones"},
        {"nombre": "Cinta aisladora",   "descripcion": "10 metros, color negro",      "precio_costo": 200,  "precio_venta": 300,  "stock_actual": 7,  "stock_minimo": 5,  "categoria": "Electricidad"},
        {"nombre": "Lija grano 120",    "descripcion": "Para madera y metal",         "precio_costo": 120,  "precio_venta": 180,  "stock_actual": 4,  "stock_minimo": 8,  "categoria": "Herramientas"},
    ]
    try:
        supabase.table("productos").insert(datos).execute()
        print("  > Datos de prueba cargados en Supabase.")
    except Exception as e:
        print(f"  ! Error al cargar seed: {e}")

seed()


# ════════════════════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════════════════════

@app.route("/api/login", methods=["POST"])
def login():
    d        = request.json or {}
    usuario  = d.get("usuario", "").strip()
    password = d.get("password", "").strip()
    u = USUARIOS.get(usuario)
    if u and u["password"] == password:
        return jsonify({"ok": True, "usuario": usuario, "rol": u["rol"]})
    return jsonify({"ok": False, "error": "Usuario o contrasena incorrectos."}), 401


# ════════════════════════════════════════════════════════════
#  FRONTEND
# ════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/logo.png")
def logo():
    return send_from_directory(".", "logo.png")


# ════════════════════════════════════════════════════════════
#  PRODUCTOS
# ════════════════════════════════════════════════════════════

@app.route("/api/productos", methods=["GET"])
def get_productos():
    result = supabase.table("productos").select("*").order("id").execute()
    for p in result.data:
        p["bajo_stock"] = p["stock_actual"] < p["stock_minimo"]
        p["precio_costo"] = float(p["precio_costo"])
        p["precio_venta"] = float(p["precio_venta"])
    return jsonify(result.data)

@app.route("/api/productos", methods=["POST"])
def crear_producto():
    d       = request.json or {}
    errores = []
    if not d.get("nombre"):    errores.append("El nombre es obligatorio.")
    if not d.get("categoria"): errores.append("La categoria es obligatoria.")
    try:    float(d.get("precio_costo", "x"))
    except: errores.append("El precio de costo debe ser un numero.")
    try:    float(d.get("precio_venta", "x"))
    except: errores.append("El precio de venta debe ser un numero.")
    try:    int(d.get("stock_actual", "x"))
    except: errores.append("El stock actual debe ser entero.")
    try:    int(d.get("stock_minimo", "x"))
    except: errores.append("El stock minimo debe ser entero.")
    if errores:
        return jsonify({"error": " | ".join(errores)}), 400

    producto = {
        "nombre": d["nombre"],
        "descripcion": d.get("descripcion", ""),
        "precio_costo": float(d["precio_costo"]),
        "precio_venta": float(d["precio_venta"]),
        "stock_actual": int(d["stock_actual"]),
        "stock_minimo": int(d["stock_minimo"]),
        "categoria": d["categoria"],
    }
    result = supabase.table("productos").insert(producto).execute()
    p = result.data[0]
    p["bajo_stock"] = p["stock_actual"] < p["stock_minimo"]
    p["precio_costo"] = float(p["precio_costo"])
    p["precio_venta"] = float(p["precio_venta"])
    return jsonify(p), 201

@app.route("/api/productos/<int:pid>/precio", methods=["PUT"])
def actualizar_precio(pid):
    d = request.json or {}
    update = {}
    if "precio_costo" in d: update["precio_costo"] = float(d["precio_costo"])
    if "precio_venta" in d: update["precio_venta"] = float(d["precio_venta"])
    if not update:
        return jsonify({"error": "Nada que actualizar."}), 400
    supabase.table("productos").update(update).eq("id", pid).execute()
    return jsonify({"ok": True})


# ════════════════════════════════════════════════════════════
#  MOVIMIENTOS (entradas de stock)
# ════════════════════════════════════════════════════════════

@app.route("/api/movimientos", methods=["GET"])
def get_movimientos():
    result = supabase.table("movimientos").select("*").order("created_at", desc=True).limit(50).execute()
    for m in result.data:
        m["precio_unitario"] = float(m["precio_unitario"])
        m["total"] = float(m["total"])
    return jsonify(result.data)

@app.route("/api/movimientos", methods=["POST"])
def registrar_movimiento():
    """Registra una ENTRADA de stock (reposicion del proveedor)."""
    d = request.json or {}
    try:
        producto_id     = int(d.get("producto_id", 0))
        cantidad        = int(d.get("cantidad", 0))
        precio_costo    = float(d.get("precio_costo", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "Datos invalidos."}), 400

    if cantidad <= 0:
        return jsonify({"error": "La cantidad debe ser mayor a cero."}), 400

    motivo = d.get("motivo", "Reposicion").strip() or "Reposicion"
    precio_venta_manual = d.get("precio_venta", None)

    # Obtener producto actual
    result = supabase.table("productos").select("*").eq("id", producto_id).execute()
    if not result.data:
        return jsonify({"error": "Producto no encontrado."}), 404
    producto = result.data[0]
    era_bajo = producto["stock_actual"] < producto["stock_minimo"]

    # Calcular nuevo stock
    nuevo_stock = producto["stock_actual"] + cantidad

    # Actualizar producto (stock + precios si se proporcionaron)
    update = {"stock_actual": nuevo_stock}
    if precio_costo > 0:
        update["precio_costo"] = precio_costo
        if precio_venta_manual is not None:
            update["precio_venta"] = float(precio_venta_manual)
        else:
            update["precio_venta"] = round(precio_costo * (1 + MARGEN_GANANCIA), 2)

    supabase.table("productos").update(update).eq("id", producto_id).execute()

    # Registrar movimiento
    precio_u = precio_costo if precio_costo > 0 else float(producto["precio_costo"])
    mov_data = {
        "producto_id": producto_id,
        "producto_nombre": producto["nombre"],
        "cantidad": cantidad,
        "tipo": "entrada",
        "motivo": motivo,
        "precio_unitario": precio_u,
        "total": round(precio_u * cantidad, 2),
    }
    mov_result = supabase.table("movimientos").insert(mov_data).execute()

    # Obtener producto actualizado
    producto_upd = supabase.table("productos").select("*").eq("id", producto_id).execute().data[0]
    producto_upd["bajo_stock"] = producto_upd["stock_actual"] < producto_upd["stock_minimo"]
    producto_upd["precio_costo"] = float(producto_upd["precio_costo"])
    producto_upd["precio_venta"] = float(producto_upd["precio_venta"])

    # Observer: notificar si stock se normalizo
    notificaciones = []
    if era_bajo and not producto_upd["bajo_stock"]:
        notificaciones = notificar_reposicion(producto_upd)

    mov = mov_result.data[0]
    mov["precio_unitario"] = float(mov["precio_unitario"])
    mov["total"] = float(mov["total"])

    return jsonify({
        "movimiento": mov,
        "producto": producto_upd,
        "notificaciones": notificaciones,
    }), 201


# ════════════════════════════════════════════════════════════
#  VENTAS (carrito -> salidas de stock)
# ════════════════════════════════════════════════════════════

@app.route("/api/ventas", methods=["POST"])
def registrar_venta():
    """Procesa el carrito completo: registra venta + movimientos de salida."""
    d = request.json or {}
    items   = d.get("items", [])
    usuario = d.get("usuario", "anon")

    if not items:
        return jsonify({"error": "El carrito esta vacio."}), 400

    # Validar todos los items primero
    productos_cache = {}
    for item in items:
        pid = int(item["producto_id"])
        cant = int(item["cantidad"])
        result = supabase.table("productos").select("*").eq("id", pid).execute()
        if not result.data:
            return jsonify({"error": f"Producto ID {pid} no encontrado."}), 404
        producto = result.data[0]
        if cant > producto["stock_actual"]:
            return jsonify({"error": f"Stock insuficiente para '{producto['nombre']}'. Stock: {producto['stock_actual']}"}), 400
        productos_cache[pid] = producto

    # Calcular total
    total_venta = sum(
        float(item["precio_unitario"]) * int(item["cantidad"])
        for item in items
    )

    # Crear registro de venta
    venta = supabase.table("ventas").insert({
        "usuario": usuario,
        "total": round(total_venta, 2),
        "items_count": len(items),
    }).execute().data[0]

    notificaciones = []

    # Procesar cada item
    for item in items:
        pid   = int(item["producto_id"])
        cant  = int(item["cantidad"])
        punit = float(item["precio_unitario"])
        sub   = round(punit * cant, 2)
        producto = productos_cache[pid]

        # Item de venta
        supabase.table("venta_items").insert({
            "venta_id": venta["id"],
            "producto_id": pid,
            "producto_nombre": producto["nombre"],
            "cantidad": cant,
            "precio_unitario": punit,
            "subtotal": sub,
        }).execute()

        # Actualizar stock
        nuevo_stock = producto["stock_actual"] - cant
        supabase.table("productos").update({"stock_actual": nuevo_stock}).eq("id", pid).execute()

        # Registrar movimiento de salida
        supabase.table("movimientos").insert({
            "producto_id": pid,
            "producto_nombre": producto["nombre"],
            "cantidad": cant,
            "tipo": "salida",
            "motivo": f"Venta #{venta['id']}",
            "precio_unitario": punit,
            "total": sub,
        }).execute()

        # Observer: verificar stock bajo
        if nuevo_stock < producto["stock_minimo"]:
            prod_updated = dict(producto)
            prod_updated["stock_actual"] = nuevo_stock
            notifs = notificar_bajo_stock(prod_updated)
            notificaciones.extend(notifs)

    venta["total"] = float(venta["total"])
    return jsonify({
        "venta": venta,
        "total": total_venta,
        "items_count": len(items),
        "notificaciones": notificaciones,
    }), 201


# ════════════════════════════════════════════════════════════
#  ALERTAS
# ════════════════════════════════════════════════════════════

@app.route("/api/alertas", methods=["GET"])
def get_alertas():
    result = supabase.table("alertas").select("*").order("created_at", desc=True).limit(50).execute()
    return jsonify(result.data)

@app.route("/api/alertas/no-leidas", methods=["GET"])
def get_alertas_no_leidas():
    result = supabase.table("alertas").select("id", count="exact").eq("leida", False).execute()
    return jsonify({"count": result.count or 0})

@app.route("/api/alertas/marcar-leidas", methods=["PUT"])
def marcar_alertas_leidas():
    supabase.table("alertas").update({"leida": True}).eq("leida", False).execute()
    return jsonify({"ok": True})


# ════════════════════════════════════════════════════════════
#  CATEGORIAS
# ════════════════════════════════════════════════════════════

@app.route("/api/categorias", methods=["GET"])
def get_categorias():
    result = supabase.table("productos").select("categoria").execute()
    categorias = sorted(set(p["categoria"] for p in result.data))
    return jsonify(categorias)


# ════════════════════════════════════════════════════════════
#  REPORTES -- JSON (Strategy)
# ════════════════════════════════════════════════════════════

@app.route("/api/reportes/<tipo>", methods=["GET"])
def get_reporte(tipo):
    if tipo == "reposicion":
        gen = GeneradorReporte(ReporteReposicion())
    elif tipo == "stock":
        gen = GeneradorReporte(ReporteStockActual())
    else:
        return jsonify({"error": "Tipo invalido."}), 400
    productos = supabase.table("productos").select("*").order("id").execute().data
    return jsonify(gen.ejecutar(productos))


# ════════════════════════════════════════════════════════════
#  EXPORTAR PDF
# ════════════════════════════════════════════════════════════

@app.route("/api/reportes/<tipo>/pdf", methods=["GET"])
def exportar_pdf(tipo):
    if tipo == "reposicion":
        gen    = GeneradorReporte(ReporteReposicion())
        titulo = "Reporte de productos a reponer"
    elif tipo == "stock":
        gen    = GeneradorReporte(ReporteStockActual())
        titulo = "Reporte de stock actual"
    else:
        return jsonify({"error": "Tipo invalido."}), 400

    productos = supabase.table("productos").select("*").order("id").execute().data
    datos = gen.ejecutar(productos)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,    bottomMargin=2*cm,
    )

    styles  = getSampleStyleSheet()
    azul    = colors.HexColor("#2563eb")
    gris_hd = colors.HexColor("#1e293b")
    gris_bg = colors.HexColor("#f1f5f9")
    rojo    = colors.HexColor("#dc2626")

    estilo_titulo = ParagraphStyle(
        "titulo", parent=styles["Normal"],
        fontSize=18, fontName="Helvetica-Bold",
        textColor=azul, spaceAfter=4, alignment=TA_CENTER,
    )
    estilo_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#6b7280"),
        alignment=TA_CENTER, spaceAfter=18,
    )

    elementos = [
        Paragraph("FerreRAP", estilo_titulo),
        Paragraph(f"{titulo} - Generado el {fecha}", estilo_sub),
    ]

    if tipo == "stock":
        cabecera = ["#", "Nombre", "Cat.", "Stock", "Min", "Costo", "Venta", "Estado"]
        col_w    = [.8*cm, 4.5*cm, 2.5*cm, 2*cm, 2*cm, 2.2*cm, 2.2*cm, 2.2*cm]
    else:
        cabecera = ["#", "Nombre", "Cat.", "Stock", "Min", "Costo", "Venta"]
        col_w    = [1*cm, 5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]

    filas = [cabecera]
    for i, p in enumerate(datos, 1):
        bajo = p.get("bajo_stock", p["stock_actual"] < p["stock_minimo"])
        fila = [
            str(i), p["nombre"], p["categoria"],
            str(p["stock_actual"]), str(p["stock_minimo"]),
            f"${p['precio_costo']:,.0f}", f"${p['precio_venta']:,.0f}",
        ]
        if tipo == "stock":
            fila.append("Reponer" if bajo else "OK")
        filas.append(fila)

    tabla = Table(filas, colWidths=col_w, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0, 0), (-1, 0), gris_hd),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUND", (0, 1), (-1, -1), [colors.white, gris_bg]),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ALIGN",         (3, 1), (-1, -1), "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
    ]
    for i, p in enumerate(datos, 1):
        bajo = p.get("bajo_stock", p["stock_actual"] < p["stock_minimo"])
        if bajo:
            ts.append(("TEXTCOLOR", (3, i), (3, i), rojo))
            ts.append(("FONTNAME",  (3, i), (3, i), "Helvetica-Bold"))
            if tipo == "stock":
                ts.append(("TEXTCOLOR", (-1, i), (-1, i), rojo))
    tabla.setStyle(TableStyle(ts))
    elementos.append(tabla)

    total    = len(datos)
    bajo_cnt = sum(1 for p in datos if p.get("bajo_stock", p["stock_actual"] < p["stock_minimo"]))
    estilo_pie = ParagraphStyle(
        "pie", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#6b7280"),
        spaceBefore=14, alignment=TA_CENTER,
    )
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph(
        f"Total: <b>{total}</b> productos  |  Bajo stock: <b><font color='#dc2626'>{bajo_cnt}</font></b>  |  FerreRAP IS2 UCP 2026",
        estilo_pie,
    ))

    doc.build(elementos)
    buffer.seek(0)
    nombre_archivo = f"ferrerap_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=nombre_archivo)


# ════════════════════════════════════════════════════════════
#  EXPORTAR EXCEL
# ════════════════════════════════════════════════════════════

@app.route("/api/reportes/<tipo>/excel", methods=["GET"])
def exportar_excel(tipo):
    if tipo == "reposicion":
        gen    = GeneradorReporte(ReporteReposicion())
        titulo = "Productos a reponer"
    elif tipo == "stock":
        gen    = GeneradorReporte(ReporteStockActual())
        titulo = "Stock actual"
    else:
        return jsonify({"error": "Tipo invalido."}), 400

    productos = supabase.table("productos").select("*").order("id").execute().data
    datos = gen.ejecutar(productos)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    azul_hd   = "1E3A5F"
    azul_sub  = "2563EB"
    gris_par  = "F1F5F9"
    rojo_hex  = "DC2626"
    verde_hex = "16A34A"
    blanco    = "FFFFFF"

    thin = Side(style="thin", color="D1D5DB")
    borde_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "FerreRAP - Sistema de Gestion de Stock"
    c.font      = Font(bold=True, size=16, color=blanco, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=azul_hd)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value     = f"{titulo}  |  Generado el {fecha}  |  IS2 UCP 2026"
    c.font      = Font(size=9, color="6B7280", italic=True, name="Calibri")
    c.fill      = PatternFill("solid", fgColor="EFF6FF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.append([])

    if tipo == "stock":
        cabecera = ["#", "Nombre", "Descripcion", "Categoria", "Stock", "Min", "Costo ($)", "Venta ($)", "Estado"]
        anchos   = [5, 22, 26, 13, 10, 10, 12, 12, 10]
    else:
        cabecera = ["#", "Nombre", "Descripcion", "Categoria", "Stock", "Min", "Costo ($)", "Venta ($)"]
        anchos   = [5, 22, 26, 13, 10, 10, 12, 12]

    fila_hd = ws.max_row + 1
    ws.append(cabecera)
    for col_i, val in enumerate(cabecera, 1):
        c = ws.cell(row=fila_hd, column=col_i)
        c.font      = Font(bold=True, color=blanco, size=10, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=azul_sub)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borde_all
    ws.row_dimensions[fila_hd].height = 20

    for i, p in enumerate(datos, 1):
        bajo = p.get("bajo_stock", p["stock_actual"] < p["stock_minimo"])
        fila = [
            i, p["nombre"], p.get("descripcion", ""), p["categoria"],
            p["stock_actual"], p["stock_minimo"],
            p["precio_costo"], p["precio_venta"],
        ]
        if tipo == "stock":
            fila.append("Reponer" if bajo else "OK")

        fila_n = ws.max_row + 1
        ws.append(fila)
        bg = gris_par if i % 2 == 0 else blanco

        for col_i, _ in enumerate(fila, 1):
            c = ws.cell(row=fila_n, column=col_i)
            c.fill   = PatternFill("solid", fgColor=bg)
            c.border = borde_all
            c.font   = Font(size=10, name="Calibri")

            if col_i == 5 and bajo:
                c.font = Font(bold=True, color=rojo_hex, size=10, name="Calibri")
            if tipo == "stock" and col_i == 9:
                c.alignment = Alignment(horizontal="center")
                c.font = Font(bold=True, color=(rojo_hex if bajo else verde_hex), size=10, name="Calibri")
            if col_i in (7, 8):
                c.number_format = '#,##0.00'
                c.alignment     = Alignment(horizontal="right")
            if col_i in (1, 5, 6):
                c.alignment = Alignment(horizontal="center")

        ws.row_dimensions[fila_n].height = 17

    for col_i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = ancho

    ws.append([])
    fila_tot = ws.max_row + 1
    total    = len(datos)
    bajo_cnt = sum(1 for p in datos if p.get("bajo_stock", p["stock_actual"] < p["stock_minimo"]))

    merge_end = "I" if tipo == "stock" else "H"
    ws.merge_cells(f"A{fila_tot}:{merge_end}{fila_tot}")
    c = ws.cell(row=fila_tot, column=1)
    c.value     = f"Total: {total} productos  |  Bajo stock: {bajo_cnt}  |  FerreRAP IS2 UCP 2026"
    c.font      = Font(italic=True, size=9, color="6B7280", name="Calibri")
    c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=fila_hd + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"ferrerap_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo,
    )


# ════════════════════════════════════════════════════════════
#  STATS
# ════════════════════════════════════════════════════════════

@app.route("/api/stats", methods=["GET"])
def get_stats():
    productos = supabase.table("productos").select("stock_actual,stock_minimo").execute().data
    alertas   = supabase.table("alertas").select("id", count="exact").execute()
    movs      = supabase.table("movimientos").select("id", count="exact").execute()
    ventas_r  = supabase.table("ventas").select("id", count="exact").execute()

    return jsonify({
        "total_productos":   len(productos),
        "bajo_stock":        sum(1 for p in productos if p["stock_actual"] < p["stock_minimo"]),
        "total_alertas":     alertas.count or 0,
        "total_movimientos": movs.count or 0,
        "total_ventas":      ventas_r.count or 0,
    })


# ════════════════════════════════════════════════════════════
#  MARGEN
# ════════════════════════════════════════════════════════════

@app.route("/api/config/margen", methods=["GET"])
def get_margen():
    return jsonify({"margen": MARGEN_GANANCIA * 100})


if __name__ == "__main__":
    print("\n  FerreRAP -- IS2 UCP 2026")
    print("  http://localhost:5000\n")
    app.run(debug=True, port=5000)
